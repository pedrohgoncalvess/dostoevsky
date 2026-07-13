import io


def extract_text(content: bytes, filename: str) -> str:
    extension = _extension(filename)

    if extension in {"txt", "md", "csv", "json", "yaml", "yml", "html", "htm", "xml", "log"}:
        return _decode_text(content)

    if extension == "pdf":
        return _extract_pdf(content)

    if extension in {"xlsx", "xls"}:
        return _extract_spreadsheet(content)

    return ""


def _extension(filename: str) -> str:
    if "." not in filename:
        return ""
    return filename.rsplit(".", 1)[-1].lower()


def _decode_text(content: bytes) -> str:
    for encoding in ("utf-8", "latin-1", "cp1252"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    return ""


def _extract_pdf(content: bytes) -> str:
    try:
        from pypdf import PdfReader
    except ImportError:
        return ""

    try:
        reader = PdfReader(io.BytesIO(content))
        parts = []
        for page in reader.pages:
            text = page.extract_text()
            if text:
                parts.append(text)
        return "\n".join(parts)
    except Exception:
        return ""


def _extract_spreadsheet(content: bytes) -> str:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return ""

    try:
        workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        parts = []
        for sheet in workbook.worksheets:
            rows = []
            for row in sheet.iter_rows(values_only=True):
                cells = [str(cell) if cell is not None else "" for cell in row]
                rows.append(" | ".join(cells))
            parts.append("\n".join(rows))
        return "\n\n".join(parts)
    except Exception:
        return ""
